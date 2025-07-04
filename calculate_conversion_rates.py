import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment

def calculate_conversion_rates(file_path):
    """
    计算D7、D14、D30、D90转化率和24小时、12小时付费转化率
    对每个注册日期的用户：
    D7转化率 = 该日期注册用户中7天内付费的人数 ÷ 该日期注册总人数
    D14转化率 = 该日期注册用户中14天内付费的人数 ÷ 该日期注册总人数
    D30转化率 = 该日期注册用户中30天内付费的人数 ÷ 该日期注册总人数
    D90转化率 = 该日期注册用户中90天内付费的人数 ÷ 该日期注册总人数
    24h转化率 = 该日期注册用户中24小时内付费的人数 ÷ 该日期注册总人数
    12h转化率 = 该日期注册用户中12小时内付费的人数 ÷ 该日期注册总人数
    """
    
    # 读取Excel文件
    print("正在读取Excel文件...")
    df = pd.read_excel(file_path)
    
    # 显示数据基本信息
    print(f"数据总行数: {len(df)}")
    print(f"列名: {df.columns.tolist()}")
    
    # 转换时间列为datetime类型
    df['注册时间'] = pd.to_datetime(df['注册时间'])
    df['首次付费时间'] = pd.to_datetime(df['首次付费时间'], errors='coerce')
    
    # 只保留有注册时间的用户
    df = df.dropna(subset=['注册时间'])
    print(f"有效注册用户数: {len(df)}")
    
    # 计算注册日期（去掉时间部分）
    df['注册日期'] = df['注册时间'].dt.date
    
    # 计算付费日期（去掉时间部分）
    df['付费日期'] = df['首次付费时间'].dt.date
    
    # 计算注册到付费的天数
    df['注册到付费天数'] = (df['首次付费时间'] - df['注册时间']).dt.days
    # 计算注册到付费的小时数
    df['注册到付费小时数'] = (df['首次付费时间'] - df['注册时间']).dt.total_seconds() / 3600
    
    # 获取每个注册日期的用户数
    daily_registrations = df.groupby('注册日期').size().reset_index(name='注册人数')
    print(f"注册日期范围: {daily_registrations['注册日期'].min()} 到 {daily_registrations['注册日期'].max()}")
    
    # 计算每个注册日期的转化率
    conversion_results = []
    
    for i, row in daily_registrations.iterrows():
        reg_date = row['注册日期']
        reg_count = row['注册人数']
        
        # 获取该日期注册的用户
        date_users = df[df['注册日期'] == reg_date]
        
        # 计算该日期注册用户的转化情况
        d12h_paid = len(date_users[(date_users['注册到付费小时数'] >= 0) & (date_users['注册到付费小时数'] <= 12)])
        d24h_paid = len(date_users[(date_users['注册到付费小时数'] >= 0) & (date_users['注册到付费小时数'] <= 24)])
        d7_paid = len(date_users[(date_users['注册到付费天数'] >= 0) & (date_users['注册到付费天数'] <= 7)])
        d14_paid = len(date_users[(date_users['注册到付费天数'] >= 0) & (date_users['注册到付费天数'] <= 14)])
        d30_paid = len(date_users[(date_users['注册到付费天数'] >= 0) & (date_users['注册到付费天数'] <= 30)])
        d90_paid = len(date_users[(date_users['注册到付费天数'] >= 0) & (date_users['注册到付费天数'] <= 90)])
        
        d12h_rate = d12h_paid / reg_count if reg_count > 0 else 0
        d24h_rate = d24h_paid / reg_count if reg_count > 0 else 0
        d7_rate = d7_paid / reg_count if reg_count > 0 else 0
        d14_rate = d14_paid / reg_count if reg_count > 0 else 0
        d30_rate = d30_paid / reg_count if reg_count > 0 else 0
        d90_rate = d90_paid / reg_count if reg_count > 0 else 0
        
        conversion_results.append({
            '注册日期': reg_date,
            '注册人数': reg_count,
            'D12h付费人数': d12h_paid,
            'D24h付费人数': d24h_paid,
            'D7付费人数': d7_paid,
            'D14付费人数': d14_paid,
            'D30付费人数': d30_paid,
            'D90付费人数': d90_paid,
            'D12h转化率': d12h_rate,
            'D24h转化率': d24h_rate,
            'D7转化率': d7_rate,
            'D14转化率': d14_rate,
            'D30转化率': d30_rate,
            'D90转化率': d90_rate
        })
    
    # 转换为DataFrame
    results_df = pd.DataFrame(conversion_results)
    
    # 计算总体平均转化率
    total_registrations = results_df['注册人数'].sum()
    total_d12h_paid = results_df['D12h付费人数'].sum()
    total_d24h_paid = results_df['D24h付费人数'].sum()
    total_d7_paid = results_df['D7付费人数'].sum()
    total_d14_paid = results_df['D14付费人数'].sum()
    total_d30_paid = results_df['D30付费人数'].sum()
    total_d90_paid = results_df['D90付费人数'].sum()
    
    overall_d12h_rate = total_d12h_paid / total_registrations if total_registrations > 0 else 0
    overall_d24h_rate = total_d24h_paid / total_registrations if total_registrations > 0 else 0
    overall_d7_rate = total_d7_paid / total_registrations if total_registrations > 0 else 0
    overall_d14_rate = total_d14_paid / total_registrations if total_registrations > 0 else 0
    overall_d30_rate = total_d30_paid / total_registrations if total_registrations > 0 else 0
    overall_d90_rate = total_d90_paid / total_registrations if total_registrations > 0 else 0
    
    print("\n=== 总体转化率 ===")
    print(f"总注册人数: {total_registrations}")
    print(f"12h内付费人数: {total_d12h_paid}")
    print(f"24h内付费人数: {total_d24h_paid}")
    print(f"总D7付费人数: {total_d7_paid}")
    print(f"总D14付费人数: {total_d14_paid}")
    print(f"总D30付费人数: {total_d30_paid}")
    print(f"总D90付费人数: {total_d90_paid}")
    print(f"12h转化率: {overall_d12h_rate:.4f} ({overall_d12h_rate*100:.2f}%)")
    print(f"24h转化率: {overall_d24h_rate:.4f} ({overall_d24h_rate*100:.2f}%)")
    print(f"总体D7转化率: {overall_d7_rate:.4f} ({overall_d7_rate*100:.2f}%)")
    print(f"总体D14转化率: {overall_d14_rate:.4f} ({overall_d14_rate*100:.2f}%)")
    print(f"总体D30转化率: {overall_d30_rate:.4f} ({overall_d30_rate*100:.2f}%)")
    print(f"总体D90转化率: {overall_d90_rate:.4f} ({overall_d90_rate*100:.2f}%)")
    
    # 显示每个注册日期的详细转化率
    print(f"\n=== 按注册日期详细转化率 ===")
    for _, row in results_df.iterrows():
        print(f"注册日期: {row['注册日期']}, 注册人数: {row['注册人数']}")
        print(f"  12h转化率: {row['D12h付费人数']}/{row['注册人数']} = {row['D12h转化率']*100:.2f}%")
        print(f"  24h转化率: {row['D24h付费人数']}/{row['注册人数']} = {row['D24h转化率']*100:.2f}%")
        print(f"  D7转化率: {row['D7付费人数']}/{row['注册人数']} = {row['D7转化率']*100:.2f}%")
        print(f"  D14转化率: {row['D14付费人数']}/{row['注册人数']} = {row['D14转化率']*100:.2f}%")
        print(f"  D30转化率: {row['D30付费人数']}/{row['注册人数']} = {row['D30转化率']*100:.2f}%")
        print(f"  D90转化率: {row['D90付费人数']}/{row['注册人数']} = {row['D90转化率']*100:.2f}%")
        print()
    
    # 生成转化率频数分布直方图数据
    print("正在生成转化率频数分布直方图...")
    
    # 创建Excel工作簿
    with pd.ExcelWriter('conversion_rates_detailed.xlsx', engine='openpyxl') as writer:
        # 写入详细数据
        results_df.to_excel(writer, sheet_name='详细数据', index=False)
        
        # 创建转化率频数分布数据
        # 定义转化率区间
        bins = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
        bin_labels = ['0-10%', '10-20%', '20-30%', '30-40%', '40-50%', '50-60%', '60-70%', '70-80%', '80-90%', '90-100%']
        
        # 计算各转化率的频数分布
        conversion_types = ['D12h转化率', 'D24h转化率', 'D7转化率', 'D14转化率', 'D30转化率', 'D90转化率']
        
        for conv_type in conversion_types:
            # 计算频数分布
            hist_data = pd.cut(results_df[conv_type], bins=bins, labels=bin_labels, include_lowest=True)
            hist_counts = hist_data.value_counts().sort_index()
            hist_freq = hist_counts / hist_counts.sum()
            hist_df = pd.DataFrame({
                '区间': bin_labels,
                '频数': hist_counts.values,
                '频率': hist_freq.values,
                '频率(%)': hist_freq.values * 100
            })
            hist_df.to_excel(writer, sheet_name=f'{conv_type}分布', index=False)
        
    # 用openpyxl添加直方图
    from openpyxl import load_workbook
    workbook = load_workbook('conversion_rates_detailed.xlsx')
    for conv_type in conversion_types:
        sheet_name = f"{conv_type}分布"
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            # 创建频数柱状图
            chart1 = BarChart()
            chart1.title = f"{conv_type}频数分布直方图"
            chart1.x_axis.title = "转化率区间"
            chart1.y_axis.title = "频数"
            data1 = Reference(ws, min_col=2, min_row=1, max_row=11)
            cats1 = Reference(ws, min_col=1, min_row=2, max_row=11)
            chart1.add_data(data1, titles_from_data=True)
            chart1.set_categories(cats1)
            ws.add_chart(chart1, "F2")
            # 创建频率柱状图
            chart2 = BarChart()
            chart2.title = f"{conv_type}频率分布直方图"
            chart2.x_axis.title = "转化率区间"
            chart2.y_axis.title = "频率(%)"
            data2 = Reference(ws, min_col=4, min_row=1, max_row=11)
            cats2 = Reference(ws, min_col=1, min_row=2, max_row=11)
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats2)
            ws.add_chart(chart2, "F15")
    workbook.save('conversion_rates_detailed.xlsx')

    # ========== 分时间段中位数趋势图 ==========
    print("正在统计指定时间段的中位数转化率并绘图...")
    period_labels = [
        "0623-0629", "0616-0622", "0609-0615", "0602-0608", "0526-0601", "0519-0525", "0512-0518", "0505-0511", "0428-0504", "0421-0427", "0414-0420"
    ]
    period_ranges = [
        ("2025-06-23", "2025-06-29"),
        ("2025-06-16", "2025-06-22"),
        ("2025-06-09", "2025-06-15"),
        ("2025-06-02", "2025-06-08"),
        ("2025-05-26", "2025-06-01"),
        ("2025-05-19", "2025-05-25"),
        ("2025-05-12", "2025-05-18"),
        ("2025-05-05", "2025-05-11"),
        ("2025-04-28", "2025-05-04"),
        ("2025-04-21", "2025-04-27"),
        ("2025-04-14", "2025-04-20")
    ]
    # 修正：将 period_ranges 的起止日期字符串转为 datetime.date 类型
    import datetime
    period_ranges = [
        (datetime.datetime.strptime(start, "%Y-%m-%d").date(), datetime.datetime.strptime(end, "%Y-%m-%d").date())
        for start, end in period_ranges
    ]
    period_conv_types = ['D12h转化率', 'D24h转化率', 'D7转化率', 'D14转化率']
    period_medians = {conv: [] for conv in period_conv_types}
    # period_labels、period_ranges逆序，需调整为时间从前到后
    period_labels = period_labels[::-1]
    period_ranges = period_ranges[::-1]
    for start, end in period_ranges:
        mask = (results_df['注册日期'] >= start) & (results_df['注册日期'] <= end)
        for conv in period_conv_types:
            median_val = results_df.loc[mask, conv].median() if mask.any() else None
            period_medians[conv].append(median_val)
    with pd.ExcelWriter('conversion_rates_detailed.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        period_df = pd.DataFrame({'时间段': period_labels})
        for conv in period_conv_types:
            period_df[conv] = [v*100 if v is not None else None for v in period_medians[conv]]
        period_df.to_excel(writer, sheet_name='分时间段中位数', index=False)
    workbook = load_workbook('conversion_rates_detailed.xlsx')
    worksheet = workbook['分时间段中位数']
    from openpyxl.chart import LineChart
    for idx, conv in enumerate(period_conv_types):
        chart = LineChart()
        chart.title = f"{conv}分时间段中位数趋势"
        chart.x_axis.title = "注册时间段"
        chart.y_axis.title = "中位数转化率(%)"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100
        data = Reference(worksheet, min_col=2+idx, min_row=1, max_row=len(period_labels)+1)
        cats = Reference(worksheet, min_col=1, min_row=2, max_row=len(period_labels)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.style = 10
        chart.height = 12
        chart.width = 18
        worksheet.add_chart(chart, f"F{2+idx*16}")
    workbook.save('conversion_rates_detailed.xlsx')
    
    return {
        'overall_d12h_rate': overall_d12h_rate,
        'overall_d24h_rate': overall_d24h_rate,
        'overall_d7_rate': overall_d7_rate,
        'overall_d14_rate': overall_d14_rate,
        'overall_d30_rate': overall_d30_rate,
        'overall_d90_rate': overall_d90_rate,
        'total_registrations': total_registrations,
        'total_d12h_paid': total_d12h_paid,
        'total_d24h_paid': total_d24h_paid,
        'total_d7_paid': total_d7_paid,
        'total_d14_paid': total_d14_paid,
        'total_d30_paid': total_d30_paid,
        'total_d90_paid': total_d90_paid,
        'detailed_results': results_df
    }

if __name__ == "__main__":
    # 计算转化率
    results = calculate_conversion_rates('Result_10.xlsx')
    
    # 保存结果到文件
    with open('conversion_rates_results.txt', 'w', encoding='utf-8') as f:
        f.write("=== 转化率计算结果 ===\n")
        f.write(f"总注册人数: {results['total_registrations']}\n")
        f.write(f"12h内付费人数: {results['total_d12h_paid']}\n")
        f.write(f"24h内付费人数: {results['total_d24h_paid']}\n")
        f.write(f"总D7付费人数: {results['total_d7_paid']}\n")
        f.write(f"总D14付费人数: {results['total_d14_paid']}\n")
        f.write(f"总D30付费人数: {results['total_d30_paid']}\n")
        f.write(f"总D90付费人数: {results['total_d90_paid']}\n")
        f.write(f"12h转化率: {results['overall_d12h_rate']:.4f} ({results['overall_d12h_rate']*100:.2f}%)\n")
        f.write(f"24h转化率: {results['overall_d24h_rate']:.4f} ({results['overall_d24h_rate']*100:.2f}%)\n")
        f.write(f"总体D7转化率: {results['overall_d7_rate']:.4f} ({results['overall_d7_rate']*100:.2f}%)\n")
        f.write(f"总体D14转化率: {results['overall_d14_rate']:.4f} ({results['overall_d14_rate']*100:.2f}%)\n")
        f.write(f"总体D30转化率: {results['overall_d30_rate']:.4f} ({results['overall_d30_rate']*100:.2f}%)\n")
        f.write(f"总体D90转化率: {results['overall_d90_rate']:.4f} ({results['overall_d90_rate']*100:.2f}%)\n")
    
    print(f"\n结果已保存到 conversion_rates_results.txt") 