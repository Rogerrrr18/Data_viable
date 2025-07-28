import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
import os
import matplotlib
matplotlib.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei']
matplotlib.rcParams['axes.unicode_minus'] = False

def calculate_conversion_rates(file_path):
    """
    计算D7、D14、D30、D90转化率和24小时、12小时付费转化率
    对每个注册日期的用户：
    D7转化率 = 该日期注册用户中7天内付费的人数 ÷ 该日期注册总人数
    D14转化率 = 该日期注册用户中14天内付费的人数 ÷ 该日期注册总人数
    D30转化率 = 该日期注册用户中30天内付费的人数 ÷ 该日期注册总人数
    D90转化率 = 该日期注册用户中90天内付费的人数 ÷ 该日期注册总人数
    24h转化率 = 该日期注册用户中24小时内付费的人数 ÷ 该日期注册总人数
    12h转化率 = 该日期注册用户中12小时内付费的人数 ÷ 该日期注册总人数
    """
    
    # 读取Excel文件
    print("正在读取Excel文件...")
    df = pd.read_excel(file_path)
    
    # 显示数据基本信息
    print(f"数据总行数: {len(df)}")
    print(f"列名: {df.columns.tolist()}")
    
    # 转换时间列为datetime类型
    df['注册时间'] = pd.to_datetime(df['注册时间'])
    df['首次付费时间'] = pd.to_datetime(df['首次付费时间'], errors='coerce')
    
    # 只保留有注册时间的用户
    df = df.dropna(subset=['注册时间'])
    print(f"有效注册用户数: {len(df)}")
    
    # 计算注册日期（去掉时间部分）
    df['注册日期'] = df['注册时间'].dt.date
    
    # 计算付费日期（去掉时间部分）
    df['付费日期'] = df['首次付费时间'].dt.date
    
    # 计算注册到付费的天数
    df['注册到付费天数'] = (df['首次付费时间'] - df['注册时间']).dt.days
    # 计算注册到付费的小时数
    df['注册到付费小时数'] = (df['首次付费时间'] - df['注册时间']).dt.total_seconds() / 3600
    
    # 获取每个注册日期的用户数
    daily_registrations = df.groupby('注册日期').size().reset_index(name='注册人数')
    print(f"注册日期范围: {daily_registrations['注册日期'].min()} 到 {daily_registrations['注册日期'].max()}")
    
    # 计算每个注册日期的转化率
    conversion_results = []
    
    for i, row in daily_registrations.iterrows():
        reg_date = row['注册日期']
        reg_count = row['注册人数']
        
        # 获取该日期注册的用户
        date_users = df[df['注册日期'] == reg_date]
        
        # 计算该日期注册用户的转化情况
        d12h_paid = len(date_users[(date_users['注册到付费小时数'] >= 0) & (date_users['注册到付费小时数'] <= 12)])
        d24h_paid = len(date_users[(date_users['注册到付费小时数'] >= 0) & (date_users['注册到付费小时数'] <= 24)])
        d7_paid = len(date_users[(date_users['注册到付费天数'] >= 0) & (date_users['注册到付费天数'] <= 7)])
        d14_paid = len(date_users[(date_users['注册到付费天数'] >= 0) & (date_users['注册到付费天数'] <= 14)])
        d30_paid = len(date_users[(date_users['注册到付费天数'] >= 0) & (date_users['注册到付费天数'] <= 30)])
        d90_paid = len(date_users[(date_users['注册到付费天数'] >= 0) & (date_users['注册到付费天数'] <= 90)])
        
        d12h_rate = d12h_paid / reg_count if reg_count > 0 else 0
        d24h_rate = d24h_paid / reg_count if reg_count > 0 else 0
        d7_rate = d7_paid / reg_count if reg_count > 0 else 0
        d14_rate = d14_paid / reg_count if reg_count > 0 else 0
        d30_rate = d30_paid / reg_count if reg_count > 0 else 0
        d90_rate = d90_paid / reg_count if reg_count > 0 else 0
        
        conversion_results.append({
            '注册日期': reg_date,
            '注册人数': reg_count,
            'D12h付费人数': d12h_paid,
            'D24h付费人数': d24h_paid,
            'D7付费人数': d7_paid,
            'D14付费人数': d14_paid,
            'D30付费人数': d30_paid,
            'D90付费人数': d90_paid,
            'D12h转化率': d12h_rate,
            'D24h转化率': d24h_rate,
            'D7转化率': d7_rate,
            'D14转化率': d14_rate,
            'D30转化率': d30_rate,
            'D90转化率': d90_rate
        })
    
    # 转换为DataFrame
    results_df = pd.DataFrame(conversion_results)
    
    # 计算总体平均转化率
    total_registrations = results_df['注册人数'].sum()
    total_d12h_paid = results_df['D12h付费人数'].sum()
    total_d24h_paid = results_df['D24h付费人数'].sum()
    total_d7_paid = results_df['D7付费人数'].sum()
    total_d14_paid = results_df['D14付费人数'].sum()
    total_d30_paid = results_df['D30付费人数'].sum()
    total_d90_paid = results_df['D90付费人数'].sum()
    
    overall_d12h_rate = total_d12h_paid / total_registrations if total_registrations > 0 else 0
    overall_d24h_rate = total_d24h_paid / total_registrations if total_registrations > 0 else 0
    overall_d7_rate = total_d7_paid / total_registrations if total_registrations > 0 else 0
    overall_d14_rate = total_d14_paid / total_registrations if total_registrations > 0 else 0
    overall_d30_rate = total_d30_paid / total_registrations if total_registrations > 0 else 0
    overall_d90_rate = total_d90_paid / total_registrations if total_registrations > 0 else 0
    
    print("\n=== 总体转化率 ===")
    print(f"总注册人数: {total_registrations}")
    print(f"12h内付费人数: {total_d12h_paid}")
    print(f"24h内付费人数: {total_d24h_paid}")
    print(f"总D7付费人数: {total_d7_paid}")
    print(f"总D14付费人数: {total_d14_paid}")
    print(f"总D30付费人数: {total_d30_paid}")
    print(f"总D90付费人数: {total_d90_paid}")
    print(f"12h转化率: {overall_d12h_rate:.4f} ({overall_d12h_rate*100:.2f}%)")
    print(f"24h转化率: {overall_d24h_rate:.4f} ({overall_d24h_rate*100:.2f}%)")
    print(f"总体D7转化率: {overall_d7_rate:.4f} ({overall_d7_rate*100:.2f}%)")
    print(f"总体D14转化率: {overall_d14_rate:.4f} ({overall_d14_rate*100:.2f}%)")
    print(f"总体D30转化率: {overall_d30_rate:.4f} ({overall_d30_rate*100:.2f}%)")
    print(f"总体D90转化率: {overall_d90_rate:.4f} ({overall_d90_rate*100:.2f}%)")
    
    # 显示每个注册日期的详细转化率
    print(f"\n=== 按注册日期详细转化率 ===")
    for _, row in results_df.iterrows():
        print(f"注册日期: {row['注册日期']}, 注册人数: {row['注册人数']}")
        print(f"  12h转化率: {row['D12h付费人数']}/{row['注册人数']} = {row['D12h转化率']*100:.2f}%")
        print(f"  24h转化率: {row['D24h付费人数']}/{row['注册人数']} = {row['D24h转化率']*100:.2f}%")
        print(f"  D7转化率: {row['D7付费人数']}/{row['注册人数']} = {row['D7转化率']*100:.2f}%")
        print(f"  D14转化率: {row['D14付费人数']}/{row['注册人数']} = {row['D14转化率']*100:.2f}%")
        print(f"  D30转化率: {row['D30付费人数']}/{row['注册人数']} = {row['D30转化率']*100:.2f}%")
        print(f"  D90转化率: {row['D90付费人数']}/{row['注册人数']} = {row['D90转化率']*100:.2f}%")
        print()
    
    # ========== 分时间段中位数趋势图 ==========
    print("正在统计指定时间段的中位数转化率并绘图...")
    period_labels = [
        "0623-0629", "0616-0622", "0609-0615", "0602-0608", "0526-0601", "0519-0525", "0512-0518", "0505-0511", "0428-0504", "0421-0427", "0414-0420"
    ]
    period_ranges = [
        ("2025-06-23", "2025-06-29"),
        ("2025-06-16", "2025-06-22"),
        ("2025-06-09", "2025-06-15"),
        ("2025-06-02", "2025-06-08"),
        ("2025-05-26", "2025-06-01"),
        ("2025-05-19", "2025-05-25"),
        ("2025-05-12", "2025-05-18"),
        ("2025-05-05", "2025-05-11"),
        ("2025-04-28", "2025-05-04"),
        ("2025-04-21", "2025-04-27"),
        ("2025-04-14", "2025-04-20")
    ]
    import datetime
    period_ranges = [
        (datetime.datetime.strptime(start, "%Y-%m-%d").date(), datetime.datetime.strptime(end, "%Y-%m-%d").date())
        for start, end in period_ranges
    ]
    # 只保留D7、D14、D30
    period_conv_types = ['D7转化率', 'D14转化率', 'D30转化率']
    period_medians = {conv: [] for conv in period_conv_types}
    period_labels = period_labels[::-1]
    period_ranges = period_ranges[::-1]
    for start, end in period_ranges:
        mask = (results_df['注册日期'] >= start) & (results_df['注册日期'] <= end)
        for conv in period_conv_types:
            median_val = results_df.loc[mask, conv].median() if mask.any() else None
            period_medians[conv].append(median_val)
    # 绘制趋势折线图并插入到Excel
    period_df = pd.DataFrame({'时间段': period_labels})
    for conv in period_conv_types:
        period_df[conv] = [v*100 if v is not None else None for v in period_medians[conv]]

    # 先保存详细数据sheet
    with pd.ExcelWriter('conversion_rates_detailed.xlsx', engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='详细数据', index=False)
        # 再保存分时间段中位数sheet
        period_df.to_excel(writer, sheet_name='分时间段中位数', index=False)

    # 绘制并保存图片
    plt.figure(figsize=(12, 6))
    for conv in period_conv_types:
        plt.plot(period_labels, [v*100 if v is not None else None for v in period_medians[conv]], marker='o', label=conv)
    plt.title('D7/D14/D30转化率分时间段中位数趋势')
    plt.xlabel('注册时间段')
    plt.ylabel('中位数转化率(%)')
    plt.legend()
    plt.grid(True, linestyle='--', alpha=0.5)
    plt.tight_layout()
    img_path = 'conversion_trend.png'
    plt.savefig(img_path)
    plt.close()

    # 将图片插入Excel
    workbook = load_workbook('conversion_rates_detailed.xlsx')
    if '趋势图' in workbook.sheetnames:
        ws = workbook['趋势图']
        for row in ws['A1:Z30']:
            for cell in row:
                cell.value = None
    else:
        ws = workbook.create_sheet('趋势图')
    img = XLImage(img_path)
    img.anchor = 'A1'
    ws.add_image(img)
    workbook.save('conversion_rates_detailed.xlsx')
    # 只返回趋势数据
    return {
        'period_labels': period_labels,
        'period_medians': period_medians
    }

if __name__ == "__main__":
    # 计算转化率
    results = calculate_conversion_rates('Result_10.xlsx')
    # 保存趋势数据到文件（可选）
    with open('conversion_rates_results.txt', 'w', encoding='utf-8') as f:
        f.write("=== 分时间段中位数趋势数据 ===\n")
        for conv, vals in results['period_medians'].items():
            f.write(f"{conv}: {vals}\n")
        f.write(f"时间段: {results['period_labels']}\n")
    print(f"\n趋势数据已保存到 conversion_rates_results.txt") 