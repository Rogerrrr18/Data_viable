import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference

# 读取原始数据
file_path = 'Result_10.xlsx'
raw_df = pd.read_excel(file_path)

# 只保留有首次付费日期的用户
paid_df = raw_df.dropna(subset=['首次付费时间']).copy()
paid_df['注册时间'] = pd.to_datetime(paid_df['注册时间'])
paid_df['注册日期'] = paid_df['注册时间'].dt.date
paid_df['首次付费时间'] = pd.to_datetime(paid_df['首次付费时间']).dt.date
paid_df['付费时长'] = (paid_df['首次付费时间'] - paid_df['注册日期']).apply(lambda x: x.days)

# 1. 频数分布直方图
bins = list(range(0, paid_df['付费时长'].max()+2))  # 0,1,2,...max+1
bin_labels = [f"{i}-{i+1}天" for i in range(0, paid_df['付费时长'].max()+1)]
hist_data = pd.cut(paid_df['付费时长'], bins=bins, labels=bin_labels, right=False, include_lowest=True)
hist_counts = hist_data.value_counts().sort_index()
hist_df = pd.DataFrame({'付费时长区间': bin_labels, '频数': hist_counts.values})

# 2. 按注册周统计中位数
paid_df['注册周'] = pd.to_datetime(paid_df['注册日期']).dt.to_period('W').apply(lambda r: r.start_time.date())
week_median = paid_df.groupby('注册周')['付费时长'].median().reset_index()
week_median = week_median.sort_values('注册周')

# 3. 平均付费时长
avg_pay_days = paid_df['付费时长'].mean()

# 4. 写入新Excel
output_file = 'pay_time_analysis.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    hist_df.to_excel(writer, sheet_name='付费时长分布', index=False)
    week_median.to_excel(writer, sheet_name='付费时长周中位数', index=False)

# 5. 用openpyxl画图
workbook = load_workbook(output_file)
# 频数分布直方图
ws = workbook['付费时长分布']
chart = BarChart()
chart.title = "付费用户注册到首次付费时长频数分布"
chart.x_axis.title = "付费时长区间"
chart.y_axis.title = "频数"
data = Reference(ws, min_col=2, min_row=1, max_row=len(hist_df)+1)
cats = Reference(ws, min_col=1, min_row=2, max_row=len(hist_df)+1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 12
chart.width = 18
ws.add_chart(chart, "E2")
# 周中位数趋势图
ws2 = workbook['付费时长周中位数']
chart2 = LineChart()
chart2.title = "每周付费时长中位数趋势"
chart2.x_axis.title = "注册周"
chart2.y_axis.title = "中位数付费时长(天)"
data2 = Reference(ws2, min_col=2, min_row=1, max_row=len(week_median)+1)
cats2 = Reference(ws2, min_col=1, min_row=2, max_row=len(week_median)+1)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.height = 12
chart2.width = 18
ws2.add_chart(chart2, "E2")
# 写入平均付费时长
ws['A' + str(len(hist_df)+3)] = f"平均付费时长: {avg_pay_days:.2f} 天"
workbook.save(output_file)
print(f"分析完成，结果已保存到 {output_file}")

# ========== 新增：1天内付费用户的付费时长（小时）分布 ==========
# 计算注册到首次付费的小时数
paid_df['付费时长_小时'] = (pd.to_datetime(paid_df['首次付费时间']) - paid_df['注册时间']).dt.total_seconds() / 3600
# 只保留1天内付费的用户
within_1day = paid_df[paid_df['付费时长_小时'] <= 24]
# 以1小时为步长分箱
hour_bins = list(range(0, 25))  # 0,1,...,24
hour_labels = [f"{i}-{i+1}h" for i in range(24)]
hist_hour_data = pd.cut(within_1day['付费时长_小时'], bins=hour_bins, labels=hour_labels, right=False, include_lowest=True)
hist_hour_counts = hist_hour_data.value_counts().sort_index()
hist_hour_df = pd.DataFrame({'付费时长区间(小时)': hour_labels, '频数': hist_hour_counts.values})
# 写入Excel
with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    hist_hour_df.to_excel(writer, sheet_name='1天内付费时长分布(小时)', index=False)
# 画图
workbook = load_workbook(output_file)
ws_hour = workbook['1天内付费时长分布(小时)']
chart_hour = BarChart()
chart_hour.title = "1天内付费用户注册到首次付费时长频数分布(小时)"
chart_hour.x_axis.title = "付费时长区间(小时)"
chart_hour.y_axis.title = "频数"
data_hour = Reference(ws_hour, min_col=2, min_row=1, max_row=len(hist_hour_df)+1)
cats_hour = Reference(ws_hour, min_col=1, min_row=2, max_row=len(hist_hour_df)+1)
chart_hour.add_data(data_hour, titles_from_data=True)
chart_hour.set_categories(cats_hour)
chart_hour.height = 12
chart_hour.width = 18
ws_hour.add_chart(chart_hour, "E2")
workbook.save(output_file)
print("1天内付费用户时长分布(小时)分析已完成！") 